#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
نرم افزار حسابداری سالن آرایش زنانه
Salon Accounting Application for Women's Beauty Salon
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import os
import sys
from datetime import datetime, timedelta, date
from collections import defaultdict
import json

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("لطفاً openpyxl را نصب کنید: pip install openpyxl")
    sys.exit(1)


# ─── Data Directory ───
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
os.makedirs(DATA_DIR, exist_ok=True)

EMPLOYEES_FILE = os.path.join(DATA_DIR, "employees.xlsx")
SERVICES_FILE = os.path.join(DATA_DIR, "services.xlsx")
TRANSACTIONS_FILE = os.path.join(DATA_DIR, "transactions.xlsx")
SETTINGS_FILE = os.path.join(DATA_DIR, "settings.json")

# ─── RTL Helper ───
RTL = "\u200f"
LRM = "\u200e"

def rtl(text):
    """Wrap text for RTL display"""
    return f"{RTL}{text}{RTL}"

def to_toman(amount):
    """Format amount in Toman with comma separators"""
    return f"{int(amount):,} {RTL}تومان"

def to_toman_short(amount):
    """Short format: e.g., 100 هزار"""
    amt = int(amount)
    if amt >= 10000:
        return f"{amt // 1000} هزار"
    return f"{amt:,}"


# ─── Color Theme ───
class Theme:
    BG = "#fce4ec"
    BG_LIGHT = "#f8bbd0"
    BG_CARD = "#ffffff"
    ACCENT = "#e91e63"
    ACCENT_LIGHT = "#f06292"
    TEXT = "#4a148c"
    TEXT_DIM = "#7b1fa2"
    SUCCESS = "#2e7d32"
    WARNING = "#e65100"
    DANGER = "#c62828"
    ENTRY_BG = "#fce4ec"
    BTN_PRIMARY = "#e91e63"
    BTN_SECONDARY = "#f8bbd0"
    BTN_HOVER = "#f06292"
    TREE_BG = "#ffffff"
    TREE_SELECT = "#f48fb1"
    TREE_ALT = "#fce4ec"


# ─── Default Data ───
DEFAULT_EMPLOYEES = [
    {"name": "مریم", "specialty": "مو", "phone": "", "share_percent": 0},
    {"name": "زهرا", "specialty": "ناخن", "phone": "", "share_percent": 0},
    {"name": "سارا", "specialty": "ابرو", "phone": "", "share_percent": 0},
    {"name": "نیلوفر", "specialty": "مژه", "phone": "", "share_percent": 0},
]

DEFAULT_SERVICES = [
    {"name": "رنگ مو", "category": "مو", "default_price": 80000},
    {"name": "مش", "category": "مو", "default_price": 120000},
    {"name": "کوتاهی مو", "category": "مو", "default_price": 50000},
    {"name": "شینیون", "category": "مو", "default_price": 100000},
    {"name": "فر دائمی", "category": "مو", "default_price": 150000},
    {"name": "ابرو برداشتن", "category": "ابرو", "default_price": 20000},
    {"name": "ابرو رنگ", "category": "ابرو", "default_price": 15000},
    {"name": "تاتو ابرو", "category": "ابرو", "default_price": 80000},
    {"name": "کاشت مژه", "category": "مژه", "default_price": 100000},
    {"name": "اکستنشن مژه", "category": "مژه", "default_price": 150000},
    {"name": "جلاس مژه", "category": "مژه", "default_price": 50000},
    {"name": "مانیکور", "category": "ناخن", "default_price": 60000},
    {"name": "پدیکور", "category": "ناخن", "default_price": 70000},
    {"name": "ژل ناخن", "category": "ناخن", "default_price": 80000},
    {"name": "کاشت ناخن", "category": "ناخن", "default_price": 120000},
    {"name": "فرچ ناخن", "category": "ناخن", "default_price": 40000},
]


# ─── Excel Manager ───
class ExcelManager:
    """Handles all Excel file operations"""

    @staticmethod
    def _ensure_workbook(filepath, headers):
        """Create workbook with headers if it doesn't exist"""
        if not os.path.exists(filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "داده‌ها"
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            wb.save(filepath)
        return load_workbook(filepath)

    @staticmethod
    def init_employees():
        wb = ExcelManager._ensure_workbook(
            EMPLOYEES_FILE,
            ["نام", "تخصص", "تلفن", "درصد سهم"]
        )
        ws = wb.active
        # Only add defaults if sheet is empty (just headers)
        if ws.max_row <= 1:
            for emp in DEFAULT_EMPLOYEES:
                ws.append([emp["name"], emp["specialty"], emp["phone"], emp["share_percent"]])
            wb.save(EMPLOYEES_FILE)

    @staticmethod
    def init_services():
        wb = ExcelManager._ensure_workbook(
            SERVICES_FILE,
            ["نام خدمت", "دسته‌بندی", "قیمت پیش‌فرض"]
        )
        ws = wb.active
        if ws.max_row <= 1:
            for svc in DEFAULT_SERVICES:
                ws.append([svc["name"], svc["category"], svc["default_price"]])
            wb.save(SERVICES_FILE)

    @staticmethod
    def init_transactions():
        ExcelManager._ensure_workbook(
            TRANSACTIONS_FILE,
            ["تاریخ", "نام مشتری", "نام خدمت", "دسته‌بندی", "نام کارمند", "مبلغ (تومان)", "یادداشت"]
        )

    @staticmethod
    def get_employees():
        wb = load_workbook(EMPLOYEES_FILE)
        ws = wb.active
        employees = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                employees.append({
                    "name": str(row[0]),
                    "specialty": str(row[1]) if row[1] else "",
                    "phone": str(row[2]) if row[2] else "",
                    "share_percent": float(row[3]) if row[3] else 0,
                })
        return employees

    @staticmethod
    def save_employees(employees):
        wb = Workbook()
        ws = wb.active
        ws.title = "داده‌ها"
        headers = ["نام", "تخصص", "تلفن", "درصد سهم"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for emp in employees:
            ws.append([emp["name"], emp["specialty"], emp["phone"], emp["share_percent"]])
        wb.save(EMPLOYEES_FILE)

    @staticmethod
    def get_services():
        wb = load_workbook(SERVICES_FILE)
        ws = wb.active
        services = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                services.append({
                    "name": str(row[0]),
                    "category": str(row[1]) if row[1] else "",
                    "default_price": int(row[2]) if row[2] else 0,
                })
        return services

    @staticmethod
    def save_services(services):
        wb = Workbook()
        ws = wb.active
        ws.title = "داده‌ها"
        headers = ["نام خدمت", "دسته‌بندی", "قیمت پیش‌فرض"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for svc in services:
            ws.append([svc["name"], svc["category"], svc["default_price"]])
        wb.save(SERVICES_FILE)

    @staticmethod
    def add_transaction(date_str, customer, service, category, employee, amount, note=""):
        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        ws.append([date_str, customer, service, category, employee, int(amount), note])
        wb.save(TRANSACTIONS_FILE)

    @staticmethod
    def get_transactions(start_date=None, end_date=None):
        """Get transactions, optionally filtered by date range"""
        if not os.path.exists(TRANSACTIONS_FILE):
            return []
        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        transactions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                t = {
                    "date": str(row[0]),
                    "customer": str(row[1]),
                    "service": str(row[2]),
                    "category": str(row[3]),
                    "employee": str(row[4]),
                    "amount": int(row[5]) if row[5] else 0,
                    "note": str(row[6]) if row[6] else "",
                }
                if start_date and end_date:
                    if start_date <= t["date"] <= end_date:
                        transactions.append(t)
                else:
                    transactions.append(t)
        return transactions

    @staticmethod
    def delete_transaction(index):
        """Delete a transaction by its row index (0-based from data rows)"""
        wb = load_workbook(TRANSACTIONS_FILE)
        ws = wb.active
        row_num = index + 2  # +2 because row 1 is header, and index is 0-based
        ws.delete_rows(row_num)
        wb.save(TRANSACTIONS_FILE)


# ─── Main Application ───
class SalonApp:
    def __init__(self):
        # Initialize data files
        ExcelManager.init_employees()
        ExcelManager.init_services()
        ExcelManager.init_transactions()

        self.root = tk.Tk()
        self.root.title("حسابداری سالن آرایش")
        self.root.geometry("1100x750")
        self.root.configure(bg=Theme.BG)
        self.root.minsize(900, 600)

        # Set RTL
        self.root.option_add("*TCombobox*Listbox.justify", "right")
        self.root.option_add("*Entry.justify", "right")

        # Configure styles
        self.setup_styles()

        # Build UI
        self.build_ui()

        # Load initial data
        self.employees = ExcelManager.get_employees()
        self.services = ExcelManager.get_services()

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use("clam")

        # General
        self.style.configure(".", background=Theme.BG, foreground=Theme.TEXT, font=("Tahoma", 11))

        # Frame
        self.style.configure("TFrame", background=Theme.BG)
        self.style.configure("Card.TFrame", background=Theme.BG_CARD)
        self.style.configure("Light.TFrame", background=Theme.BG_LIGHT)

        # Label
        self.style.configure("TLabel", background=Theme.BG, foreground=Theme.TEXT, font=("Tahoma", 11))
        self.style.configure("Title.TLabel", font=("Tahoma", 18, "bold"), foreground=Theme.ACCENT, background=Theme.BG)
        self.style.configure("Subtitle.TLabel", font=("Tahoma", 13, "bold"), foreground=Theme.TEXT, background=Theme.BG)
        self.style.configure("Card.TLabel", background=Theme.BG_CARD, foreground=Theme.TEXT, font=("Tahoma", 11))
        self.style.configure("CardTitle.TLabel", background=Theme.BG_CARD, foreground=Theme.ACCENT, font=("Tahoma", 12, "bold"))
        self.style.configure("Big.TLabel", font=("Tahoma", 28, "bold"), foreground=Theme.SUCCESS, background=Theme.BG_CARD)
        self.style.configure("Dim.TLabel", background=Theme.BG, foreground=Theme.TEXT_DIM, font=("Tahoma", 9))

        # Button
        self.style.configure("Accent.TButton",
                           background=Theme.ACCENT, foreground=Theme.TEXT,
                           font=("Tahoma", 11, "bold"), padding=(15, 8))
        self.style.map("Accent.TButton",
                      background=[("active", Theme.BTN_HOVER), ("pressed", Theme.DANGER)])

        self.style.configure("Secondary.TButton",
                           background=Theme.BTN_SECONDARY, foreground=Theme.TEXT,
                           font=("Tahoma", 11), padding=(12, 6))
        self.style.map("Secondary.TButton",
                      background=[("active", Theme.ACCENT)])

        self.style.configure("Success.TButton",
                           background=Theme.SUCCESS, foreground=Theme.TEXT,
                           font=("Tahoma", 11, "bold"), padding=(15, 8))
        self.style.map("Success.TButton",
                      background=[("active", "#27ae60")])

        self.style.configure("Danger.TButton",
                           background=Theme.DANGER, foreground=Theme.TEXT,
                           font=("Tahoma", 10), padding=(10, 5))
        self.style.map("Danger.TButton",
                      background=[("active", "#c0392b")])

        # Treeview
        self.style.configure("Treeview",
                           background=Theme.TREE_BG,
                           foreground=Theme.TEXT,
                           fieldbackground=Theme.TREE_BG,
                           font=("Tahoma", 11),
                           rowheight=35,
                           borderwidth=0)
        self.style.configure("Treeview.Heading",
                           background=Theme.BG_CARD,
                           foreground=Theme.ACCENT,
                           font=("Tahoma", 11, "bold"),
                           relief="flat")
        self.style.map("Treeview",
                      background=[("selected", Theme.TREE_SELECT)],
                      foreground=[("selected", Theme.TEXT)])

        # Notebook (tabs)
        self.style.configure("TNotebook", background=Theme.BG, borderwidth=0)
        self.style.configure("TNotebook.Tab",
                           background=Theme.BG_LIGHT,
                           foreground=Theme.TEXT_DIM,
                           font=("Tahoma", 12),
                           padding=(20, 8))
        self.style.map("TNotebook.Tab",
                      background=[("selected", Theme.BG_CARD)],
                      foreground=[("selected", Theme.ACCENT)])

        # Combobox
        self.style.configure("TCombobox", fieldbackground=Theme.ENTRY_BG,
                           background=Theme.ENTRY_BG, foreground=Theme.TEXT,
                           arrowcolor=Theme.ACCENT, font=("Tahoma", 11))
        self.style.map("TCombobox",
                      fieldbackground=[("readonly", Theme.ENTRY_BG)])

        # Entry
        self.style.configure("TEntry", fieldbackground=Theme.ENTRY_BG,
                           foreground=Theme.TEXT, insertcolor=Theme.ACCENT)

        # Spinbox
        self.style.configure("TSpinbox", fieldbackground=Theme.ENTRY_BG,
                           foreground=Theme.TEXT, arrowcolor=Theme.ACCENT)

    def build_ui(self):
        # Header
        header = ttk.Frame(self.root)
        header.pack(fill="x", padx=20, pady=(15, 5))
        ttk.Label(header, text="💇‍♀️ حسابداری سالن آرایش", style="Title.TLabel").pack(side="right")
        ttk.Label(header, text=f"📅 امروز: {PersianDate.today_str()}", style="Dim.TLabel").pack(side="left")

        # Notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=15, pady=10)

        # Tabs
        self.tab_main = ttk.Frame(self.notebook)
        self.tab_employees = ttk.Frame(self.notebook)
        self.tab_services = ttk.Frame(self.notebook)
        self.tab_reports = ttk.Frame(self.notebook)
        self.tab_monthly = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_main, text="  💰 ثبت تراکنش  ")
        self.notebook.add(self.tab_reports, text="  📊 گزارش امروز  ")
        self.notebook.add(self.tab_monthly, text="  📅 گزارش ماهانه  ")
        self.notebook.add(self.tab_employees, text="  👥 کارمندان  ")
        self.notebook.add(self.tab_services, text="  🛎️ خدمات  ")

        # Build each tab
        self.build_main_tab()
        self.build_reports_tab()
        self.build_monthly_tab()
        self.build_employees_tab()
        self.build_services_tab()

    # ─── Main Transaction Tab ───
    def build_main_tab(self):
        # Left: Transaction form
        left = ttk.Frame(self.tab_main)
        left.pack(side="right", fill="both", expand=True, padx=(10, 5), pady=10)

        # Customer info
        customer_frame = ttk.LabelFrame(left, text=" اطلاعات مشتری ", style="Card.TFrame")
        customer_frame.pack(fill="x", pady=(0, 10))

        row1 = ttk.Frame(customer_frame, style="Card.TFrame")
        row1.pack(fill="x", padx=10, pady=10)

        ttk.Label(row1, text="نام مشتری:", style="Card.TLabel").pack(side="right", padx=(0, 5))
        self.customer_name = tk.Entry(row1, font=("Tahoma", 12), bg=Theme.ENTRY_BG,
                                      fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                      justify="right", width=25, relief="flat")
        self.customer_name.pack(side="right", padx=5)

        ttk.Label(row1, text="📱 تلفن:", style="Card.TLabel").pack(side="right", padx=(20, 5))
        self.customer_phone = tk.Entry(row1, font=("Tahoma", 12), bg=Theme.ENTRY_BG,
                                       fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                       justify="right", width=18, relief="flat")
        self.customer_phone.pack(side="right", padx=5)

        # Transaction items
        items_frame = ttk.LabelFrame(left, text=" خدمات انجام شده ", style="Card.TFrame")
        items_frame.pack(fill="both", expand=True, pady=(0, 10))

        # Items list (Treeview)
        cols = ("service", "category", "employee", "amount", "commission")
        self.items_tree = ttk.Treeview(items_frame, columns=cols, show="headings", height=5)
        self.items_tree.heading("service", text="خدمت", anchor="e")
        self.items_tree.heading("category", text="دسته", anchor="e")
        self.items_tree.heading("employee", text="کارمند", anchor="e")
        self.items_tree.heading("amount", text="مبلغ", anchor="e")
        self.items_tree.heading("commission", text="پورسانت", anchor="e")
        self.items_tree.column("service", width=150, anchor="e")
        self.items_tree.column("category", width=80, anchor="e")
        self.items_tree.column("employee", width=100, anchor="e")
        self.items_tree.column("amount", width=110, anchor="e")
        self.items_tree.column("commission", width=100, anchor="e")

        scrollbar = ttk.Scrollbar(items_frame, orient="vertical", command=self.items_tree.yview)
        self.items_tree.configure(yscrollcommand=scrollbar.set)

        self.items_tree.pack(side="right", fill="both", expand=True, padx=(10, 0), pady=5)
        scrollbar.pack(side="left", fill="y", pady=5, padx=(10, 0))

        # Add item form
        add_frame = ttk.Frame(items_frame, style="Card.TFrame")
        add_frame.pack(fill="x", padx=10, pady=(0, 5))

        # Row for adding items
        add_row = ttk.Frame(add_frame, style="Card.TFrame")
        add_row.pack(fill="x")

        ttk.Label(add_row, text="خدمت:", style="Card.TLabel").pack(side="right", padx=(0, 3))
        self.item_service = ttk.Combobox(add_row, width=18, font=("Tahoma", 11), state="readonly")
        self.item_service.pack(side="right", padx=3)

        ttk.Label(add_row, text="کارمند:", style="Card.TLabel").pack(side="right", padx=(10, 3))
        self.item_employee = ttk.Combobox(add_row, width=12, font=("Tahoma", 11), state="readonly")
        self.item_employee.pack(side="right", padx=3)

        ttk.Label(add_row, text="مبلغ (تومان):", style="Card.TLabel").pack(side="right", padx=(10, 3))
        self.item_amount = tk.Entry(add_row, font=("Tahoma", 11), bg=Theme.ENTRY_BG,
                                    fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                    justify="right", width=12, relief="flat")
        self.item_amount.pack(side="right", padx=3)

        add_btn = ttk.Button(add_row, text="➕ افزودن", style="Accent.TButton",
                           command=self.add_item)
        add_btn.pack(side="right", padx=10)

        remove_btn = ttk.Button(add_row, text="❌ حذف", style="Danger.TButton",
                              command=self.remove_item)
        remove_btn.pack(side="right", padx=3)

        # Total and submit
        bottom = ttk.Frame(left, style="Card.TFrame")
        bottom.pack(fill="x")

        self.total_var = tk.StringVar(value="جمع کل: ۰ تومان")
        ttk.Label(bottom, textvariable=self.total_var, style="CardTitle.TLabel",
                 font=("Tahoma", 14, "bold")).pack(side="right", padx=20, pady=10)

        submit_btn = ttk.Button(bottom, text="✅ ثبت نهایی", style="Success.TButton",
                              command=self.submit_transaction)
        submit_btn.pack(side="left", padx=15, pady=10)

        clear_btn = ttk.Button(bottom, text="🗑️ پاک کردن", style="Secondary.TButton",
                             command=self.clear_transaction)
        clear_btn.pack(side="left", padx=5, pady=10)

        # Items data
        self.transaction_items = []

        # Bind service selection to auto-fill price
        self.item_service.bind("<<ComboboxSelected>>", self.on_service_selected)

        # Populate comboboxes
        self.refresh_comboboxes()

    def refresh_comboboxes(self):
        self.employees = ExcelManager.get_employees()
        self.services = ExcelManager.get_services()

        emp_names = [f"{e['name']} ({e['specialty']})" for e in self.employees]
        svc_names = [f"{s['name']} - {to_toman_short(s['default_price'])}" for s in self.services]

        self.item_employee["values"] = emp_names
        self.item_service["values"] = svc_names

    def on_service_selected(self, event=None):
        idx = self.item_service.current()
        if idx >= 0:
            svc = self.services[idx]
            self.item_amount.delete(0, tk.END)
            self.item_amount.insert(0, str(svc["default_price"]))

    def add_item(self):
        svc_idx = self.item_service.current()
        emp_idx = self.item_employee.current()
        amount = self.item_amount.get().strip()

        if svc_idx < 0:
            messagebox.showwarning("خطا", "لطفاً خدمت را انتخاب کنید")
            return
        if emp_idx < 0:
            messagebox.showwarning("خطا", "لطفاً کارمند را انتخاب کنید")
            return
        if not amount or not amount.isdigit():
            messagebox.showwarning("خطا", "لطفاً مبلغ صحیح وارد کنید")
            return

        svc = self.services[svc_idx]
        emp = self.employees[emp_idx]
        commission = int(int(amount) * emp["share_percent"] / 100)

        self.transaction_items.append({
            "service": svc["name"],
            "category": svc["category"],
            "employee": emp["name"],
            "amount": int(amount),
            "commission": commission,
        })

        self.items_tree.insert("", "end",
            values=(svc["name"], svc["category"], emp["name"], f"{int(amount):,}", f"{commission:,}"))

        self.update_total()
        self.item_amount.delete(0, tk.END)

    def remove_item(self):
        selected = self.items_tree.selection()
        if not selected:
            return
        idx = self.items_tree.index(selected[0])
        self.items_tree.delete(selected[0])
        self.transaction_items.pop(idx)
        self.update_total()

    def update_total(self):
        total = sum(item["amount"] for item in self.transaction_items)
        self.total_var.set(f"جمع کل: {total:,} تومان")

    def submit_transaction(self):
        customer = self.customer_name.get().strip()
        if not customer:
            messagebox.showwarning("خطا", "لطفاً نام مشتری را وارد کنید")
            return
        if not self.transaction_items:
            messagebox.showwarning("خطا", "حداقل یک خدمت اضافه کنید")
            return

        today = PersianDate.today_str()
        note = self.customer_phone.get().strip()

        for item in self.transaction_items:
            ExcelManager.add_transaction(
                date_str=today,
                customer=customer,
                service=item["service"],
                category=item["category"],
                employee=item["employee"],
                amount=item["amount"],
                note=note
            )

        total = sum(item["amount"] for item in self.transaction_items)
        messagebox.showinfo("موفق", f"✅ تراکنش ثبت شد!\n\nمشتری: {customer}\nتعداد خدمات: {len(self.transaction_items)}\nجمع کل: {total:,} تومان")

        self.clear_transaction()
        self.refresh_reports()

    def clear_transaction(self):
        self.customer_name.delete(0, tk.END)
        self.customer_phone.delete(0, tk.END)
        self.items_tree.delete(*self.items_tree.get_children())
        self.transaction_items = []
        self.total_var.set("جمع کل: ۰ تومان")

    # ─── Reports Tab ───
    def build_reports_tab(self):
        # Top bar
        top = ttk.Frame(self.tab_reports)
        top.pack(fill="x", padx=15, pady=10)

        ttk.Label(top, text="📊 گزارش امروز", style="Subtitle.TLabel").pack(side="right")

        refresh_btn = ttk.Button(top, text="🔄 بروزرسانی", style="Secondary.TButton",
                               command=self.refresh_reports)
        refresh_btn.pack(side="left", padx=5)

        # Summary cards
        cards = ttk.Frame(self.tab_reports)
        cards.pack(fill="x", padx=15, pady=(0, 10))

        self.card_total = self.create_card(cards, "💰 جمع کل امروز", "۰")
        self.card_count = self.create_card(cards, "👥 تعداد مشتری", "۰")
        self.card_services = self.create_card(cards, "🛎️ تعداد خدمات", "۰")
        self.card_commission = self.create_card(cards, "💸 پورسانت کارمندان", "۰", color=Theme.WARNING)

        # Transaction table
        table_frame = ttk.Frame(self.tab_reports)
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        cols = ("time", "customer", "service", "category", "employee", "amount", "commission")
        self.report_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=12)
        self.report_tree.heading("time", text="تاریخ", anchor="e")
        self.report_tree.heading("customer", text="مشتری", anchor="e")
        self.report_tree.heading("service", text="خدمت", anchor="e")
        self.report_tree.heading("category", text="دسته", anchor="e")
        self.report_tree.heading("employee", text="کارمند", anchor="e")
        self.report_tree.heading("amount", text="مبلغ", anchor="e")
        self.report_tree.heading("commission", text="پورسانت", anchor="e")
        self.report_tree.column("time", width=90, anchor="e")
        self.report_tree.column("customer", width=120, anchor="e")
        self.report_tree.column("service", width=110, anchor="e")
        self.report_tree.column("category", width=70, anchor="e")
        self.report_tree.column("employee", width=90, anchor="e")
        self.report_tree.column("amount", width=100, anchor="e")
        self.report_tree.column("commission", width=100, anchor="e")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.report_tree.yview)
        self.report_tree.configure(yscrollcommand=scrollbar.set)
        self.report_tree.pack(side="right", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

    def create_card(self, parent, title, value, color=None):
        card = tk.Frame(parent, bg=Theme.BG_CARD, highlightbackground=Theme.ACCENT,
                       highlightthickness=1)
        card.pack(side="right", fill="both", expand=True, padx=5)

        tk.Label(card, text=title, bg=Theme.BG_CARD, fg=Theme.TEXT_DIM,
                font=("Tahoma", 10), anchor="e").pack(pady=(8, 0), padx=10, anchor="e")

        var = tk.StringVar(value=value)
        tk.Label(card, textvariable=var, bg=Theme.BG_CARD, fg=color or Theme.SUCCESS,
                font=("Tahoma", 22, "bold"), anchor="e").pack(pady=(0, 8), padx=10, anchor="e")

        return {"card": card, "var": var}

    def refresh_reports(self):
        today = PersianDate.today_str()
        transactions = ExcelManager.get_transactions(start_date=today, end_date=today)

        # Build employee share lookup
        emp_shares = {e["name"]: e["share_percent"] for e in self.employees}

        # Update cards
        total = sum(t["amount"] for t in transactions)
        customers = len(set(t["customer"] for t in transactions))
        services_count = len(transactions)
        total_commission = sum(int(t["amount"] * emp_shares.get(t["employee"], 0) / 100) for t in transactions)

        self.card_total["var"].set(f"{total:,}")
        self.card_count["var"].set(str(customers))
        self.card_services["var"].set(str(services_count))
        self.card_commission["var"].set(f"{total_commission:,}")

        # Update table
        self.report_tree.delete(*self.report_tree.get_children())
        for t in reversed(transactions):
            comm = int(t["amount"] * emp_shares.get(t["employee"], 0) / 100)
            self.report_tree.insert("", "end",
                values=(t["date"], t["customer"], t["service"],
                       t["category"], t["employee"], f"{t['amount']:,}", f"{comm:,}"))

    # ─── Monthly Report Tab ───
    def build_monthly_tab(self):
        top = ttk.Frame(self.tab_monthly)
        top.pack(fill="x", padx=15, pady=10)

        ttk.Label(top, text="📅 گزارش ماهانه بر اساس کارمند", style="Subtitle.TLabel").pack(side="right")

        # Month selector
        ttk.Label(top, text="ماه:", style="TLabel").pack(side="left", padx=(20, 5))
        now = datetime.now()
        self.month_var = tk.StringVar(value=f"{now.year}/{now.month:02d}")
        months = []
        for y in range(1400, 1410):
            for m in range(1, 13):
                months.append(f"{y}/{m:02d}")
        self.month_combo = ttk.Combobox(top, textvariable=self.month_var, values=months,
                                        width=12, font=("Tahoma", 11), state="readonly")
        self.month_combo.pack(side="left", padx=5)

        refresh_btn = ttk.Button(top, text="🔄 نمایش", style="Secondary.TButton",
                               command=self.refresh_monthly)
        refresh_btn.pack(side="left", padx=5)

        export_btn = ttk.Button(top, text="📥 خروجی اکسل", style="Accent.TButton",
                              command=self.export_monthly)
        export_btn.pack(side="left", padx=5)

        # Employee summary cards area
        self.monthly_cards_frame = ttk.Frame(self.tab_monthly)
        self.monthly_cards_frame.pack(fill="x", padx=15, pady=10)

        # Detail table
        table_frame = ttk.Frame(self.tab_monthly)
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        cols = ("employee", "category", "total_services", "total_amount", "commission")
        self.monthly_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=10)
        self.monthly_tree.heading("employee", text="کارمند", anchor="e")
        self.monthly_tree.heading("category", text="دسته خدمت", anchor="e")
        self.monthly_tree.heading("total_services", text="تعداد", anchor="e")
        self.monthly_tree.heading("total_amount", text="جمع درآمد", anchor="e")
        self.monthly_tree.heading("commission", text="پورسانت", anchor="e")
        self.monthly_tree.column("employee", width=150, anchor="e")
        self.monthly_tree.column("category", width=120, anchor="e")
        self.monthly_tree.column("total_services", width=80, anchor="e")
        self.monthly_tree.column("total_amount", width=130, anchor="e")
        self.monthly_tree.column("commission", width=130, anchor="e")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.monthly_tree.yview)
        self.monthly_tree.configure(yscrollcommand=scrollbar.set)
        self.monthly_tree.pack(side="right", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

    def refresh_monthly(self):
        month_str = self.month_var.get()
        if not month_str:
            return

        # Parse month to get date range
        try:
            parts = month_str.split("/")
            year = int(parts[0])
            month = int(parts[1])
        except:
            messagebox.showerror("خطا", "فرمت ماه صحیح نیست")
            return

        # Get start and end of month in Shamsi
        start_date = f"{year}/{month:02d}/01"
        if month < 12:
            end_date = f"{year}/{month+1:02d}/01"
        else:
            end_date = f"{year+1}/01/01"

        # Get all transactions and filter by date prefix
        # Get all transactions and filter by date prefix
        all_transactions = ExcelManager.get_transactions()
        monthly = [t for t in all_transactions if t["date"] >= start_date and t["date"] < end_date]
        # Build employee share lookup
        emp_shares = {e["name"]: e["share_percent"] for e in self.employees}

        # Clear previous
        for widget in self.monthly_cards_frame.winfo_children():
            widget.destroy()

        # Create employee summary cards
        emp_totals = defaultdict(lambda: {"count": 0, "amount": 0, "commission": 0})
        for t in monthly:
            comm = int(t["amount"] * emp_shares.get(t["employee"], 0) / 100)
            emp_totals[t["employee"]]["count"] += 1
            emp_totals[t["employee"]]["amount"] += t["amount"]
            emp_totals[t["employee"]]["commission"] += comm

        for emp_name, data in sorted(emp_totals.items(), key=lambda x: -x[1]["amount"]):
            card = tk.Frame(self.monthly_cards_frame, bg=Theme.BG_CARD,
                          highlightbackground=Theme.ACCENT, highlightthickness=1)
            card.pack(side="right", fill="both", expand=True, padx=5, pady=5)

            tk.Label(card, text=f"👤 {emp_name}", bg=Theme.BG_CARD, fg=Theme.TEXT,
                    font=("Tahoma", 12, "bold"), anchor="e").pack(pady=(8, 2), padx=10, anchor="e")
            tk.Label(card, text=f"{data['count']} خدمت", bg=Theme.BG_CARD, fg=Theme.TEXT_DIM,
                    font=("Tahoma", 10)).pack(anchor="e", padx=10)
            tk.Label(card, text=f"{data['amount']:,} تومان", bg=Theme.BG_CARD, fg=Theme.SUCCESS,
                    font=("Tahoma", 16, "bold")).pack(pady=(2, 2), anchor="center", padx=10)
            if data["commission"] > 0:
                tk.Label(card, text=f"💸 پورسانت: {data['commission']:,} تومان", bg=Theme.BG_CARD, fg=Theme.WARNING,
                        font=("Tahoma", 11, "bold")).pack(pady=(0, 8), anchor="center", padx=10)

        # Detail table
        self.monthly_tree.delete(*self.monthly_tree.get_children())
        emp_cat = defaultdict(lambda: defaultdict(lambda: {"count": 0, "amount": 0, "commission": 0}))
        for t in monthly:
            comm = int(t["amount"] * emp_shares.get(t["employee"], 0) / 100)
            emp_cat[t["employee"]][t["category"]]["count"] += 1
            emp_cat[t["employee"]][t["category"]]["amount"] += t["amount"]
            emp_cat[t["employee"]][t["category"]]["commission"] += comm

        for emp_name in sorted(emp_cat.keys()):
            for cat_name in sorted(emp_cat[emp_name].keys()):
                data = emp_cat[emp_name][cat_name]
                self.monthly_tree.insert("", "end",
                    values=(emp_name, cat_name, data["count"], f"{data['amount']:,}", f"{data['commission']:,}"))

        # Grand total row
        grand_total = sum(t["amount"] for t in monthly)
        grand_commission = sum(int(t["amount"] * emp_shares.get(t["employee"], 0) / 100) for t in monthly)
        self.monthly_tree.insert("", "end",
            values=("مجموع کل", "", len(monthly), f"{grand_total:,}", f"{grand_commission:,}"),
            tags=("total",))
        self.monthly_tree.tag_configure("total", background=Theme.ACCENT, foreground=Theme.TEXT)

    def export_monthly(self):
        month_str = self.month_var.get()
        try:
            parts = month_str.split("/")
            year = int(parts[0])
            month = int(parts[1])
        except:
            messagebox.showerror("خطا", "فرمت ماه صحیح نیست")
            return

        start_date = f"{year}/{month:02d}/01"
        end_date = f"{year}/{month+1:02d}/01" if month < 12 else f"{year+1}/01/01"

        all_transactions = ExcelManager.get_transactions()
        monthly = [t for t in all_transactions if start_date <= t["date"] < end_date]

        if not monthly:
            messagebox.showinfo("اطلاع", "تراکنشی برای این ماه وجود ندارد")
            return

        # Create Excel report
        wb = Workbook()
        ws = wb.active
        ws.title = f"گزارش {month_str}"

        # Header
        headers = ["تاریخ", "مشتری", "خدمت", "دسته", "کارمند", "مبلغ", "پورسانت", "یادداشت"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="E94560", end_color="E94560", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Build employee share lookup
        emp_shares = {e["name"]: e["share_percent"] for e in self.employees}

        # Data
        for row, t in enumerate(monthly, 2):
            ws.cell(row=row, column=1, value=t["date"])
            ws.cell(row=row, column=2, value=t["customer"])
            ws.cell(row=row, column=3, value=t["service"])
            ws.cell(row=row, column=4, value=t["category"])
            ws.cell(row=row, column=5, value=t["employee"])
            ws.cell(row=row, column=6, value=t["amount"])
            comm = int(t["amount"] * emp_shares.get(t["employee"], 0) / 100)
            ws.cell(row=row, column=7, value=comm)
            ws.cell(row=row, column=8, value=t["note"])

        # Auto-size columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

        filepath = os.path.join(DATA_DIR, f"report_{year}_{month:02d}.xlsx")
        wb.save(filepath)
        messagebox.showinfo("موفق", f"📥 گزارش ذخیره شد:\n{filepath}")

    # ─── Employees Tab ───
    def build_employees_tab(self):
        top = ttk.Frame(self.tab_employees)
        top.pack(fill="x", padx=15, pady=10)

        ttk.Label(top, text="👥 مدیریت کارمندان", style="Subtitle.TLabel").pack(side="right")

        add_btn = ttk.Button(top, text="➕ افزودن کارمند", style="Accent.TButton",
                           command=self.add_employee)
        add_btn.pack(side="left", padx=5)

        edit_btn = ttk.Button(top, text="✏️ ویرایش", style="Secondary.TButton",
                            command=self.edit_employee)
        edit_btn.pack(side="left", padx=5)

        del_btn = ttk.Button(top, text="❌ حذف", style="Danger.TButton",
                           command=self.delete_employee)
        del_btn.pack(side="left", padx=5)

        # Table
        table_frame = ttk.Frame(self.tab_employees)
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        cols = ("name", "specialty", "phone", "share")
        self.emp_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=15)
        self.emp_tree.heading("name", text="نام", anchor="e")
        self.emp_tree.heading("specialty", text="تخصص", anchor="e")
        self.emp_tree.heading("phone", text="تلفن", anchor="e")
        self.emp_tree.heading("share", text="درصد سهم", anchor="e")
        self.emp_tree.column("name", width=200, anchor="e")
        self.emp_tree.column("specialty", width=150, anchor="e")
        self.emp_tree.column("phone", width=150, anchor="e")
        self.emp_tree.column("share", width=100, anchor="e")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.emp_tree.yview)
        self.emp_tree.configure(yscrollcommand=scrollbar.set)
        self.emp_tree.pack(side="right", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

        self.refresh_employee_table()

    def refresh_employee_table(self):
        self.employees = ExcelManager.get_employees()
        self.emp_tree.delete(*self.emp_tree.get_children())
        for emp in self.employees:
            self.emp_tree.insert("", "end",
                values=(emp["name"], emp["specialty"], emp["phone"], f"{emp['share_percent']}%"))

    def add_employee(self):
        dialog = EmployeeDialog(self.root, "افزودن کارمند")
        if dialog.result:
            self.employees.append(dialog.result)
            ExcelManager.save_employees(self.employees)
            self.refresh_employee_table()
            self.refresh_comboboxes()

    def edit_employee(self):
        selected = self.emp_tree.selection()
        if not selected:
            messagebox.showwarning("خطا", "کارمندی انتخاب نشده")
            return
        idx = self.emp_tree.index(selected[0])
        dialog = EmployeeDialog(self.root, "ویرایش کارمند", self.employees[idx])
        if dialog.result:
            self.employees[idx] = dialog.result
            ExcelManager.save_employees(self.employees)
            self.refresh_employee_table()
            self.refresh_comboboxes()

    def delete_employee(self):
        selected = self.emp_tree.selection()
        if not selected:
            messagebox.showwarning("خطا", "کارمندی انتخاب نشده")
            return
        idx = self.emp_tree.index(selected[0])
        name = self.employees[idx]["name"]
        if messagebox.askyesno("تأیید", f"آیا از حذف '{name}' اطمینان دارید؟"):
            self.employees.pop(idx)
            ExcelManager.save_employees(self.employees)
            self.refresh_employee_table()
            self.refresh_comboboxes()

    # ─── Services Tab ───
    def build_services_tab(self):
        top = ttk.Frame(self.tab_services)
        top.pack(fill="x", padx=15, pady=10)

        ttk.Label(top, text="🛎️ مدیریت خدمات", style="Subtitle.TLabel").pack(side="right")

        add_btn = ttk.Button(top, text="➕ افزودن خدمت", style="Accent.TButton",
                           command=self.add_service)
        add_btn.pack(side="left", padx=5)

        edit_btn = ttk.Button(top, text="✏️ ویرایش", style="Secondary.TButton",
                            command=self.edit_service)
        edit_btn.pack(side="left", padx=5)

        del_btn = ttk.Button(top, text="❌ حذف", style="Danger.TButton",
                           command=self.delete_service)
        del_btn.pack(side="left", padx=5)

        # Table
        table_frame = ttk.Frame(self.tab_services)
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        cols = ("name", "category", "price")
        self.svc_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=15)
        self.svc_tree.heading("name", text="نام خدمت", anchor="e")
        self.svc_tree.heading("category", text="دسته‌بندی", anchor="e")
        self.svc_tree.heading("price", text="قیمت پیش‌فرض", anchor="e")
        self.svc_tree.column("name", width=250, anchor="e")
        self.svc_tree.column("category", width=150, anchor="e")
        self.svc_tree.column("price", width=150, anchor="e")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.svc_tree.yview)
        self.svc_tree.configure(yscrollcommand=scrollbar.set)
        self.svc_tree.pack(side="right", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

        self.refresh_service_table()

    def refresh_service_table(self):
        self.services = ExcelManager.get_services()
        self.svc_tree.delete(*self.svc_tree.get_children())
        for svc in self.services:
            self.svc_tree.insert("", "end",
                values=(svc["name"], svc["category"], f"{svc['default_price']:,}"))

    def add_service(self):
        dialog = ServiceDialog(self.root, "افزودن خدمت")
        if dialog.result:
            self.services.append(dialog.result)
            ExcelManager.save_services(self.services)
            self.refresh_service_table()
            self.refresh_comboboxes()

    def edit_service(self):
        selected = self.svc_tree.selection()
        if not selected:
            messagebox.showwarning("خطا", "خدمتی انتخاب نشده")
            return
        idx = self.svc_tree.index(selected[0])
        dialog = ServiceDialog(self.root, "ویرایش خدمت", self.services[idx])
        if dialog.result:
            self.services[idx] = dialog.result
            ExcelManager.save_services(self.services)
            self.refresh_service_table()
            self.refresh_comboboxes()

    def delete_service(self):
        selected = self.svc_tree.selection()
        if not selected:
            messagebox.showwarning("خطا", "خدمتی انتخاب نشده")
            return
        idx = self.svc_tree.index(selected[0])
        name = self.services[idx]["name"]
        if messagebox.askyesno("تأیید", f"آیا از حذف '{name}' اطمینان دارید؟"):
            self.services.pop(idx)
            ExcelManager.save_services(self.services)
            self.refresh_service_table()
            self.refresh_comboboxes()

    def run(self):
        self.root.mainloop()


# ─── Persian Date Helper ───
class PersianDate:
    """Simple Shamsi (Jalali) date converter"""

    @staticmethod
    def gregorian_to_jalali(gy, gm, gd):
        """Convert Gregorian date to Jalali (Shamsi)"""
        g_d_m = [0, 31, 59, 90, 120, 151, 181, 212, 243, 273, 304, 334]
        if gm > 2:
            gy2 = gy + 1
        else:
            gy2 = gy
        days = 355666 + (365 * gy) + ((gy2 + 3) // 4) - ((gy2 + 99) // 100) + \
               ((gy2 + 399) // 400) + gd + g_d_m[gm - 1]
        jy = -1595 + (33 * (days // 12053))
        days %= 12053
        jy += 4 * (days // 1461)
        days %= 1461
        if days > 365:
            jy += (days - 1) // 365
            days = (days - 1) % 365
        if days < 186:
            jm = 1 + (days // 31)
            jd = 1 + (days % 31)
        else:
            jm = 7 + ((days - 186) // 30)
            jd = 1 + ((days - 186) % 30)
        return jy, jm, jd

    @staticmethod
    def today_str():
        """Return today's date as Jalali string: YYYY/MM/DD"""
        now = datetime.now()
        jy, jm, jd = PersianDate.gregorian_to_jalali(now.year, now.month, now.day)
        return f"{jy}/{jm:02d}/{jd:02d}"


# ─── Dialogs ───
class EmployeeDialog:
    def __init__(self, parent, title, employee=None):
        self.result = None
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("400x300")
        self.dialog.configure(bg=Theme.BG)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog)
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        ttk.Label(frame, text="نام:", style="TLabel").pack(anchor="e", pady=(0, 3))
        self.name_entry = tk.Entry(frame, font=("Tahoma", 12), bg=Theme.ENTRY_BG,
                                  fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                  justify="right", relief="flat")
        self.name_entry.pack(fill="x", pady=(0, 10))

        ttk.Label(frame, text="تخصص:", style="TLabel").pack(anchor="e", pady=(0, 3))
        self.specialty_var = tk.StringVar()
        self.specialty_combo = ttk.Combobox(frame, textvariable=self.specialty_var,
                                            values=["مو", "ناخن", "ابرو", "مژه", "صورت"],
                                            font=("Tahoma", 11), state="readonly")
        self.specialty_combo.pack(fill="x", pady=(0, 10))

        ttk.Label(frame, text="تلفن:", style="TLabel").pack(anchor="e", pady=(0, 3))
        self.phone_entry = tk.Entry(frame, font=("Tahoma", 12), bg=Theme.ENTRY_BG,
                                   fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                   justify="right", relief="flat")
        self.phone_entry.pack(fill="x", pady=(0, 10))

        ttk.Label(frame, text="درصد سهم:", style="TLabel").pack(anchor="e", pady=(0, 3))
        self.share_entry = tk.Entry(frame, font=("Tahoma", 12), bg=Theme.ENTRY_BG,
                                   fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                   justify="right", relief="flat")
        self.share_entry.pack(fill="x", pady=(0, 10))

        # Pre-fill if editing
        if employee:
            self.name_entry.insert(0, employee["name"])
            self.specialty_var.set(employee["specialty"])
            self.phone_entry.insert(0, employee["phone"])
            self.share_entry.insert(0, str(employee["share_percent"]))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x", pady=(10, 0))

        ttk.Button(btn_frame, text="ذخیره", style="Success.TButton",
                  command=self.save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="لغو", style="Secondary.TButton",
                  command=self.cancel).pack(side="left", padx=5)

        self.dialog.wait_window()

    def save(self):
        name = self.name_entry.get().strip()
        specialty = self.specialty_var.get()
        phone = self.phone_entry.get().strip()
        share = self.share_entry.get().strip() or "0"

        if not name:
            messagebox.showwarning("خطا", "نام الزامی است")
            return

        try:
            share_val = float(share)
        except ValueError:
            messagebox.showwarning("خطا", "درصد سهم باید عدد باشد")
            return

        self.result = {
            "name": name,
            "specialty": specialty,
            "phone": phone,
            "share_percent": share_val,
        }
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


class ServiceDialog:
    def __init__(self, parent, title, service=None):
        self.result = None
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("400x250")
        self.dialog.configure(bg=Theme.BG)
        self.dialog.transient(parent)
        self.dialog.grab_set()

        frame = ttk.Frame(self.dialog)
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        ttk.Label(frame, text="نام خدمت:", style="TLabel").pack(anchor="e", pady=(0, 3))
        self.name_entry = tk.Entry(frame, font=("Tahoma", 12), bg=Theme.ENTRY_BG,
                                  fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                  justify="right", relief="flat")
        self.name_entry.pack(fill="x", pady=(0, 10))

        ttk.Label(frame, text="دسته‌بندی:", style="TLabel").pack(anchor="e", pady=(0, 3))
        self.category_var = tk.StringVar()
        self.category_combo = ttk.Combobox(frame, textvariable=self.category_var,
                                           values=["مو", "ناخن", "ابرو", "مژه", "صورت"],
                                           font=("Tahoma", 11), state="readonly")
        self.category_combo.pack(fill="x", pady=(0, 10))

        ttk.Label(frame, text="قیمت پیش‌فرض (تومان):", style="TLabel").pack(anchor="e", pady=(0, 3))
        self.price_entry = tk.Entry(frame, font=("Tahoma", 12), bg=Theme.ENTRY_BG,
                                   fg=Theme.TEXT, insertbackground=Theme.ACCENT,
                                   justify="right", relief="flat")
        self.price_entry.pack(fill="x", pady=(0, 10))

        if service:
            self.name_entry.insert(0, service["name"])
            self.category_var.set(service["category"])
            self.price_entry.insert(0, str(service["default_price"]))

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x", pady=(10, 0))

        ttk.Button(btn_frame, text="ذخیره", style="Success.TButton",
                  command=self.save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="لغو", style="Secondary.TButton",
                  command=self.cancel).pack(side="left", padx=5)

        self.dialog.wait_window()

    def save(self):
        name = self.name_entry.get().strip()
        category = self.category_var.get()
        price = self.price_entry.get().strip()

        if not name:
            messagebox.showwarning("خطا", "نام خدمت الزامی است")
            return
        if not price or not price.isdigit():
            messagebox.showwarning("خطا", "قیمت باید عدد صحیح باشد")
            return

        self.result = {
            "name": name,
            "category": category,
            "default_price": int(price),
        }
        self.dialog.destroy()

    def cancel(self):
        self.dialog.destroy()


# ─── Run ───
if __name__ == "__main__":
    app = SalonApp()
    app.run()
